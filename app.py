import streamlit as st
import pandas as pd
from io import BytesIO
import tempfile
import shutil
from openpyxl import load_workbook
import datetime as dt


# ─── 공통 로직: 노임단가 & 손해보험요율 로드 ───
@st.cache_data
def load_노임단가():
    url = "https://docs.google.com/spreadsheets/d/e/2PACX-1vSlIUPyOxmtCRrXFqQKZ7Ge3um3xi5VCaua1OvyC27Y7vw5jqJhzbFpnTeb-fcxGS3_wNxuhnBddRl4/pub?output=csv"
    df = pd.read_csv(url)
    df.columns = [c.strip() for c in df.columns]
    return df

@st.cache_data
def load_손해보험요율():
    url = "https://docs.google.com/spreadsheets/d/e/2PACX-1vRzdleYSG38-1FpxjoIkQbhWHbwY4himRBCO7LR8wWkCg0bnplhSTecIHNInZ-5NCjcjfuwmGotRFd_/pub?output=csv"
    df = pd.read_csv(url)
    return df

# ─── 조경 전용 처리 ───
def run_조경():
    st.title("🌿 조경 설계 대가 산출 프로그램")

    직급리스트 = ["기술사", "특급기술자", "고급기술자", "중급기술자", "초급기술자"]
    # 조경 전용 글로벌 데이터
    난이도_map = {
        "도시공원": [
            "단순 (소공원·묘지공원·보행자 전용도로·광장·도시공원 내 시설 교체사업)",
            "보통 (국가도시공원·근린공원·체육공원·수변공원·도시농업공원·유원지·공공공지·광장(재생사업))",
            "복잡1 (어린이공원·문화공원·역사공원·방재공원)",
            "복잡2 (도시공원(재생사업))",
        ],
        "공동주택 및 대지의 조경": [
            "보통 (공동주택 조경)",
            "복잡1 (주택정원·건축물 조경·옥상조경(옥상정원))",
            "복잡2 (실내조경(실내정원))",
        ],
        "녹지 및 도시숲": [
            "단순 (완충 녹지·가로변 녹지·가로수·경관숲)",
            "보통 (연결 녹지·경관 녹지·유휴지 녹화·마을숲·유아숲체험원)",
            "복잡 (가로변 녹지(정원형)·학교숲·도시숲)",
        ],
        "주제형 사업": [
            "단순 (야영장·둘레길·하천 경관 개선, 생태통로, 숲길조성)",
            "보통 (테마시설 조성·관광지·관광지 활성화 사업·가로 환경개선 등)",
            "복잡 (관광단지·동물원·골프장·스키장·2종 이상 복합 사업)",
        ],
    }

    성격_coeffs = {
        "도시공원":               1.0,
        "공동주택 및 대지의 조경": 1.1,
        "녹지 및 도시숲":          0.8,
        "주제형 사업":             1.2,
    }

    난이도_coeffs = {
        "도시공원": {"단순": 0.9, "보통": 1.0, "복잡1": 1.1, "복잡2": 1.2},
        "공동주택 및 대지의 조경": {"보통": 1.0, "복잡1": 1.1, "복잡2": 1.2},
        "녹지 및 도시숲": {"단순": 0.9, "보통": 1.0, "복잡": 1.1},
        "주제형 사업": {"단순": 0.9, "보통": 1.0, "복잡": 1.1},
    }


    @st.cache_data
    def load_조경_기준(설계유형):
        urls = {
            "기본설계": "https://docs.google.com/spreadsheets/d/e/2PACX-1vSffous-aCPOAcKkizEiELMpZVECskizIxlP2Vn_eHTfLnviFFCn0S1fAZPy0OkFLE508TspBu9VuuV/pub?output=csv",
            "실시설계": "https://docs.google.com/spreadsheets/d/e/2PACX-1vSRBhcxu6BMlio-obyGAj44PhEP07BPAFC9l53gad1TqZPgQyAkj289qqshKNFQ1jHYYtIrWlO9wKOm/pub?output=csv",
            "기본 및 실시설계": "https://docs.google.com/spreadsheets/d/e/2PACX-1vTcmEUxkny-pnOAPFvb67DH-MpINOZY6PqCGz9m6U3DUzFcTeqgd7Mvm7Ss1_m7i0RYE4locXoE1HuK/pub?output=csv"
        }
        return pd.read_csv(urls.get(설계유형, urls["기본설계"]))

    # 조경 전용 템플릿 엑셀 생성 함수
    def build_조경_excel(template_path="template_조경.xlsx") -> BytesIO:
        # 1) 템플릿 복사
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp.close()
        try:
            shutil.copy(template_path, tmp.name)
        except FileNotFoundError:
            st.error(f"템플릿 파일을 찾을 수 없습니다: {template_path}")
            return None

        # 2) 갑지 시트에 기본 정보 채우기
        wb = load_workbook(tmp.name)
        ws = wb["갑지"]
        ws["D10"].value = st.session_state.get("용역명", "")
        ws["G22"].value = st.session_state.get("발주기관명", "")
        raw = st.session_state.get("도급예정액", 0)
        ws["G20"].value = f"{int(raw // 1000) * 1000:,} 원"
        ws["A1"].value = dt.date.today().strftime("%Y-%m-%d")
        wb.save(tmp.name)

        # 3) 나머지 시트(overwrite) – pandas ExcelWriter 사용
        with pd.ExcelWriter(tmp.name, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            # A) 내역서
            df_detail = st.session_state.get("df_detail", pd.DataFrame())
            if not df_detail.empty:
                df_detail.to_excel(writer,
                                sheet_name="내역서",
                                index=False,
                                header=False,
                                startrow=2)

            # B) 투입인원 및 내역
            df_person = st.session_state.get("투입인원DF", pd.DataFrame())
            if not df_person.empty:
                df_person.to_excel(writer,
                                sheet_name="투입인원 및 내역",
                                index=False,
                                header=False,
                                startrow=2)

            # C) 투입인원수 산정기준
            df_basis = st.session_state.get("기준계산결과", pd.DataFrame())
            if not df_basis.empty:
                df_basis.to_excel(writer,
                                sheet_name="투입인원수 산정기준",
                                index=False,
                                header=False,
                                startrow=2)

            # D) 노임단가
            df_wage = st.session_state.get("최종_단가", pd.DataFrame())
            if not df_wage.empty:
                df_wage.to_excel(writer,
                                sheet_name="노임단가",
                                index=False,
                                header=False,
                                startrow=2)

            # E) 손해보험요율
            df_ins = st.session_state.get("보험요율DF", pd.DataFrame())
            if not df_ins.empty:
                df_ins.to_excel(writer,
                                sheet_name="손해보험요율",
                                index=False,
                                header=False,
                                startrow=2)

        # 4) 최종 BytesIO에 담아서 반환
        buf = BytesIO()
        with open(tmp.name, "rb") as f:
            buf.write(f.read())
        buf.seek(0)
        return buf
    (
        tab_기초입력,
        tab_갑지,
        tab_내역서,
        tab_투입인원및내역,
        tab_산정기준,
        tab_노임단가,
        tab_손해보험요율
    ) = st.tabs([
        "기초입력",
        "갑지",
        "내역서",
        "투입인원 및 내역",
        "투입인원수 산정기준",
        "노임단가",
        "보험요율"
    ])

    with tab_기초입력:
        st.header("기초입력")
        용역명 = st.text_input("용역명", value=st.session_state.get("용역명", "")) 
        st.session_state["용역명"] = 용역명
        발주기관명 = st.text_input("발주기관명", value=st.session_state.get("발주기관명", ""))
        st.session_state["발주기관명"] = 발주기관명
        
        options = ["기본설계", "실시설계", "기본 및 실시설계"]
        current = st.session_state.get("설계유형", "기본설계")
        index = options.index(current) if current in options else 0
        설계유형 = st.radio(
            "설계유형을 선택하세요",
            options,
            index=index,
            key="설계유형_radio"
        )
        st.session_state["설계유형"] = 설계유형    
       
        면적 = st.number_input("대상 면적 (㎡)",
                        min_value=100.0, step=100.0,
                        value=st.session_state.get("면적",100.0))
        st.session_state["면적"] = 면적

        성격_options = [
            "도시공원",
            "공동주택 및 대지의 조경",
            "녹지 및 도시숲",
            "주제형 사업"
        ]
        default_성격 = st.session_state.get("대상지_성격", "도시공원")
        if default_성격 not in 성격_options:                 
                default_성격 = "도시공원"

        대상지_성격 = st.selectbox(
           "대상지 성격",
           성격_options,
        index=성격_options.index(default_성격)
        )
        st.session_state["대상지_성격"] = 대상지_성격

        options_nd = 난이도_map.get(대상지_성격, ["단순","보통","복잡"])
        prev_nd    = st.session_state.get("난이도", options_nd[0])
        if prev_nd not in options_nd:
            prev_nd = options_nd[0]

        난이도 = st.selectbox(
            "업무 난이도",
            options_nd,
            index=options_nd.index(prev_nd),
            key="난이도"
        )

        전단계_활용 = st.checkbox(
            "기본계획 등 설계에 활용할 전 단계 성과물이 있습니까?", 
            value=False
        )
        st.session_state["전단계_활용"] = 전단계_활용

        if st.button("🔄  입력값 모두 초기화", help="용역명·면적 등 기초입력과 계산 결과를 지웁니다."):
            reset_keys = [
                "용역명", "발주기관명",
                "선택공종", "설계유형",
                "면적", "대상지_성격", "난이도", "전단계_활용",
                "기준계산결과", "직접인건비", "도급예정액",
            ]
            reset_keys += [k for k in st.session_state if k.startswith("기간_")]

            for k in reset_keys:
                st.session_state.pop(k, None)    

            st.rerun()

    with tab_갑지:
        import datetime
        today = datetime.date.today().strftime("%Y-%m-%d")

        st.markdown(f"##### 날짜: {today}")

        st.markdown(
            f"<h2 style='text-align:center;'>{용역명}</h2>",
            unsafe_allow_html=True
        )

        if "도급예정액" not in st.session_state:
            st.info("먼저 ‘내역서’ 탭에서 **산출 완료✅** 버튼을 눌러 금액을 확정하세요.")
        else:
            raw        = st.session_state["도급예정액"]
            용역비      = int(raw // 1000) * 1000   
            st.write(f"**용역비:** {용역비:,.0f} 원")

        발주기관 = st.session_state.get("발주기관명", "")
        st.write(f"**발주기관:** {발주기관}")

        if "도급예정액" in st.session_state and st.session_state["도급예정액"] > 0:
            excel_buf = build_조경_excel("template.xlsx")
            st.download_button(
                label="⬇️ 산출내역서(Excel) 다운로드",
                data=excel_buf,
                file_name=f"{st.session_state['용역명']}_조경설계 내역서.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.caption("※ 산출 완료 후 버튼이 활성화됩니다.")

    with tab_내역서:
        st.header("내역서")
        st.caption("※ 각 숫자를 수정한 뒤 **Enter** 를 눌러야 계산이 반영됩니다.")

        직접인건비 = st.session_state.get("직접인건비")
        if 직접인건비 is None:
            st.warning("먼저 ‘투입인원 및 내역’ 탭에서 직접인건비를 계산해 주세요.")
        else:
            제경비율   = st.number_input("제경비율 (110~120%)",     value=110.0, step=0.1)
            직접경비   = st.number_input("직접경비 금액 (원)", value=5_000_000, step=1_000)
            기술료율   = st.number_input("기술료율 (20~40%)",     value=20.0, step=0.1)
            공제율    = st.number_input("손해공제비율 (관람집회공사,0.432적용)",   value=0.432, step=0.001)
            부가세율   = st.number_input("부가가치세율 (%)",   value=10.0, step=0.1)

            제경비     = 직접인건비 * 제경비율   / 100
            기술료     = (직접인건비 + 제경비) * 기술료율 / 100
            손해공제비 = (직접인건비 + 제경비 + 직접경비 + 기술료) * 공제율   / 100
            부가세     = (직접인건비 + 제경비 + 직접경비 + 기술료 + 손해공제비) * 부가세율 / 100
            도급예정액  = 직접인건비 + 제경비 + 직접경비 + 기술료 + 손해공제비 + 부가세

            if st.button("✅ 산출 완료"):
                st.session_state["도급예정액"] = 도급예정액
                st.success(f"도급예정액 {도급예정액:,.0f}원이 저장되었습니다.")
            else:
                st.info("▶️ 값이 맞다면 ‘✅ 산출 완료’ 버튼을 눌러 주세요.")

            rows = [
                {"공종":"직접인건비", "규격":"-",         "수량":"-", "단위":"", "총액":직접인건비, "노무비":직접인건비, "경비":"",        "비고":""},
                {"공종":"제경비",     "규격":"직접인건비×율", "수량":"-", "단위":"", "총액":"",    "노무비":"",        "경비":제경비,    "비고":f"{제경비율}%"},
                {"공종":"직접경비",   "규격":"제출도서 인쇄",   "수량":1,   "단위":"식", "총액":"",    "노무비":"",        "경비":직접경비,  "비고":""},
                {"공종":"기술료",     "규격":"(직접인건비+제경비)×율","수량":"-", "단위":"", "총액":"",    "노무비":"",        "경비":기술료,    "비고":f"{기술료율}%"},
                {"공종":"손해공제비", "규격":"용역비×율",      "수량":"-", "단위":"", "총액":"",    "노무비":"",        "경비":손해공제비,"비고":f"{공제율}"},
                {"공종":"부가가치세", "규격":"합계×율",       "수량":"-", "단위":"", "총액":"",    "노무비":"",        "경비":부가세,    "비고":f"{부가세율}%"},
                {"공종":"도급예정액","규격":"",             "수량":"-", "단위":"", "총액":도급예정액,"노무비":"",        "경비":"",        "비고":""},
            ]
            df = pd.DataFrame(rows)

            for c in ["총액","노무비","경비"]:
                df[c] = df[c].apply(lambda x: f"{int(x):,}" if isinstance(x,(int,float)) else x)

            st.dataframe(df[[
                "공종","규격","수량","단위",
                "총액","노무비","경비","비고"
            ]])
            st.session_state["df_detail"] = df

    with tab_투입인원및내역:
        st.header("투입인원 및 내역")

        기준결과    = st.session_state.get("기준계산결과")
        노임단가_df = st.session_state.get("최종_단가")

        if 기준결과 is None or 노임단가_df is None:
            st.warning("먼저 '투입인원수 산정기준'과 '노임단가' 탭을 완료해주세요.")
        else:
            결과표 = 기준결과.copy()
            결과표 = 결과표[결과표["단위"] != ""].reset_index(drop=True)
            for 직급 in 직급리스트:
                결과표[직급] = pd.to_numeric(결과표[직급], errors='coerce').fillna(0.0)

            횟수_키워드 = ["위원회 심의", "주민설명회", "관계기관 협의"]
            기간값 = {}
            n = len(결과표)
            half = (n + 1) // 2
            left, right = st.columns(2)

            with left:
                for idx, row in 결과표.iloc[:half].iterrows():
                    업무 = row["업무구분"]
                    단위 = row["단위"].strip()  # 단위 칼럼 읽어오기

                    if 단위 == "식":
                        기본값 = 1
                        라벨 = f"{업무} (식)"

                    elif any(kw in 업무 for kw in 횟수_키워드):
                        if "주민설명회" in 업무:
                            기본값 = 2
                        else:
                            기본값 = 1
                        라벨 = f"{업무} (회)"

                    else:
                        기본값 = 2
                        라벨 = f"{업무} 기간 (일)"

                    값 = st.number_input(
                    라벨,
                    min_value=0,
                    step=1,
                    value=int(st.session_state.get(f"기간_{idx}", 기본값)),
                    key=f"기간_=L_{idx}"
                    )
                    기간값[idx] = 값

            with right:
                for idx, row in 결과표.iloc[half:].iterrows():
                    업무 = row["업무구분"]
                    단위 = row["단위"].strip()  
                    if 단위 == "식":
                        기본값 = 1
                        라벨 = f"{업무} (식)"
                    else:
                        기본값 = 2
                        라벨 = f"{업무} 기간 (일)"

                    값 = st.number_input(
                    라벨,
                    min_value=0,
                    step=1,
                    value=int(st.session_state.get(f"기간_{idx}", 기본값)),
                    key=f"기간_=L_{idx}"
                    )
                    기간값[idx] = 값

            결과표["기간"] = [기간값[i] for i in range(n)]

            노임단가_df.columns   = [c.strip() for c in 노임단가_df.columns]
            노임단가_df["직종명"] = 노임단가_df["직종명"].astype(str).str.strip()
            노임단가_df["건설"]   = (
                노임단가_df["건설"]
                .astype(str)
                .str.replace(",", "")
                .str.strip()
                .astype(float)
            )
            직급리스트 = ["기술사","특급기술자","고급기술자","중급기술자","초급기술자"]
            건설단가 = {}
            for 직급 in 직급리스트:
                sub = 노임단가_df[노임단가_df["직종명"] == 직급]
                건설단가[직급] = float(sub["건설"].iloc[0]) if not sub.empty else 0.0


            계산된_계 = []
            for _, row in 결과표.iterrows():
                인건비합 = sum(row[직급] * 건설단가[직급] for 직급 in 직급리스트)
                계산된_계.append(round(인건비합 * row["기간"], 2))
            결과표["계"] = 계산된_계

            표시열   = ["업무구분","계"] + 직급리스트 + ["기간"]
            sum_계   = 결과표["계"].sum()
            총계행    = {c: "" for c in 표시열}
            총계행["업무구분"] = "총계"
            총계행["계"]    = sum_계
            total_df = pd.DataFrame([총계행])
            final_df = pd.concat([total_df, 결과표[표시열]], ignore_index=True)

            for c in ["계","기간"] + 직급리스트:
                final_df[c] = final_df[c].apply(
                    lambda x: f"{x:,.2f}" if isinstance(x, (int, float)) else x
                )

            sum_계 = 결과표["계"].sum()

            st.session_state["직접인건비"] = sum_계

            st.session_state["투입인원DF"] = final_df

            st.subheader("📊 기술자별 투입 인원 및 총액")
            st.dataframe(final_df)

    with tab_산정기준:
        st.header("투입인원수 산정기준")

        공종        = st.session_state.get("선택공종")
        설계유형     = st.session_state.get("설계유형")
        대상_면적    = st.session_state.get("면적", 0)
        성격        = st.session_state.get("대상지_성격")
        전단계_활용  = st.session_state.get("전단계_활용", False)

        if 설계유형 in ["기본설계", "실시설계", "기본 및 실시설계"]:
            기준표 = load_조경_기준(설계유형).copy()
            for 직급 in 직급리스트:
                기준표[직급] = pd.to_numeric(기준표[직급], errors="coerce").fillna(0.0)
            if 면적 <= 5000:
                환산계수 = (대상_면적 / 5000) ** 0.7
            else:
                환산계수 = (대상_면적 / 5000) ** 0.4
            full_nd_label = st.session_state.get("난이도", "")
            diff_key = full_nd_label.split()[0] if full_nd_label else ""
            a2 = 성격_coeffs.get(성격, 1.0)
            a3 = 난이도_coeffs.get(성격, {}).get(diff_key, 1.0)

            for 직급 in 직급리스트:
                계산값, 계산식 = [], []
                for _, row in 기준표.iterrows():
                    base = row[직급]
                    v = base
                    parts = []

                    if row["환산계수(α₁)"] == "적용":
                        v *= 환산계수; parts.append(f"{환산계수:.3f}")

                    if row["보정계수(α₂, α₃)"] == "적용" and row["업무구분"] not in ["조사", "기술협의"]:
                        v *= a2 * a3
                        parts.append(f"{a2:.3f}")  # 성격계수
                        parts.append(f"{a3:.3f}")  # 난이도계수


                    if 전단계_활용:
                        first_token = str(row["업무구분"]).strip().split()[0]
                        if first_token.startswith("2.1"):        
                            v *= 0.7
                            parts.append("0.700")

                    formula = f"{base} × " + " × ".join(parts) if parts else f"{base} (고정)"
                    계산값.append(round(v,2)); 계산식.append(formula)
                기준표[직급] = 계산값
                기준표[f"{직급}_계산식"] = 계산식

            기준표 = 기준표.fillna("") 

            for 직급 in 직급리스트:
                calc_col = f"{직급}_계산식"
                if calc_col in 기준표.columns:
                    기준표[calc_col] = 기준표[calc_col].str.replace(r"\s*\(고정\)", "", regex=True)
            mask = 기준표['단위'] == ""
            cols = 직급리스트 + [f"{j}_계산식" for j in 직급리스트]
            기준표.loc[mask, cols] = ""

            st.subheader(f"📊 {설계유형} 계산된 투입인원 (인·일)")
            표시열 = ["업무구분", "단위"] + sum([[j, f"{j}_계산식"] for j in 직급리스트], [])
            st.dataframe(기준표[표시열])
            st.session_state["기준계산결과"] = 기준표

        else:
            st.info("‘기본설계’, ‘실시설계’, ‘기본 및 실시설계’ 중 하나를 모두 선택해야 계산이 표시됩니다.")

    with tab_노임단가:
        st.header("노임단가")
        df_wage = load_노임단가()
        st.dataframe(df_wage)
        st.session_state["최종_단가"] = df_wage

    with tab_손해보험요율:
        st.header("손해보험요율")
        st.dataframe(load_손해보험요율())
        st.session_state["보험요율DF"] = load_손해보험요율()

# ─── 환경영향평가 전용 처리 ───
def run_환경영향평가대행():
    st.title("🌱 환경영향평가 대행 비용 산출 프로그램")

    def build_환경_excel(template_path="template_env.xlsx") -> BytesIO:
        """
        'template_env.xlsx' 복사 후
        1) ‘갑지’ 시트: 날짜, 용역명, 발주기관명, 용역비(도급예정액) 입력
        2) 나머지 시트(내역서, 투입인원 및 내역, 투입인원수 산정기준, 노임단가, 손해보험요율)를
           pandas DataFrame으로 덮어쓰기
        """
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp.close()
        try:
            shutil.copy(template_path, tmp.name)
        except FileNotFoundError:
            st.error(f"템플릿 파일을 찾을 수 없습니다: {template_path}")
            return None

        wb = load_workbook(tmp.name)
        ws = wb["갑지"]

        ws["A1"].value = dt.date.today().strftime("%Y-%m-%d")
        용역명_env = st.session_state.get("용역명_env", "")
        ws["D10"].value = 용역명_env
        발주기관명_env = st.session_state.get("발주기관명_env", "")
        ws["G22"].value = 발주기관명_env
        raw_env = st.session_state.get("도급예정액_env", 0)
        용역비_env = int(raw_env // 1000) * 1000
        ws["G20"].value = f"{용역비_env:,} 원"

        wb.save(tmp.name)

        with pd.ExcelWriter(tmp.name, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            df_detail_env = st.session_state.get("df_detail_env", pd.DataFrame())
            if not df_detail_env.empty:
                df_detail_env.to_excel(
                    writer,
                    sheet_name="내역서",
                    index=False,
                    header=False,
                    startrow=2
                )

            df_person_env = st.session_state.get("투입인원DF_env", pd.DataFrame())
            if not df_person_env.empty:
                df_person_env.to_excel(
                    writer,
                    sheet_name="투입인원 및 내역",
                    index=False,
                    header=False,
                    startrow=2
                )

            df_basis_env = st.session_state.get("기준결과_env", pd.DataFrame())
            if not df_basis_env.empty:
                cols_to_drop = [
                    "환산계수",
                    "α₁(환산계수)",
                    "보정계수(가)",
                    "보정계수(나)",
                    "보정계수(다)",
                    "보정계수(라)",
                ]
                drop_list = [c for c in cols_to_drop if c in df_basis_env.columns]
                df_for_excel_env = df_basis_env.drop(columns=drop_list)
                df_for_excel_env.to_excel(
                    writer,
                    sheet_name="투입인원수 산정기준",
                    index=False,
                    header=False,
                    startrow=2
                )

            df_wage_env = st.session_state.get("최종_단가_env", pd.DataFrame())
            if not df_wage_env.empty:
                df_wage_env.to_excel(
                    writer,
                    sheet_name="노임단가",
                    index=False,
                    header=False,
                    startrow=2
                )

            df_ins_env = st.session_state.get("보험요율DF_env", pd.DataFrame())
            if not df_ins_env.empty:
                df_ins_env.to_excel(
                    writer,
                    sheet_name="손해보험요율",
                    index=False,
                    header=False,
                    startrow=2
                )

        buf = BytesIO()
        with open(tmp.name, "rb") as f:
            buf.write(f.read())
        buf.seek(0)
        return buf

    직급리스트 = ["기술사", "특급기술자", "고급기술자", "중급기술자", "초급기술자"]

    @st.cache_data
    def load_env_basis(설계유형_env):
        urls = {
            "소규모 환경영향평가 대행": (
                "https://docs.google.com/spreadsheets/d/e/2PACX-1vQK25xZ-K2mo1mpr_Sz4cUEv8lHW6gY0-Ps0BVW-GSFRk7_53WBVHXBIBOQChTrtdYaJIP8T1p1U0sX/pub?output=csv"
            ),
        }
        url = urls.get(설계유형_env, urls["소규모 환경영향평가 대행"])
        df = pd.read_csv(url)
        df.columns = [c.strip() for c in df.columns]
        for 직급 in 직급리스트:
            if 직급 in df.columns:
                df[직급] = pd.to_numeric(df[직급], errors="coerce").fillna(0.0)
        return df

    (
        tab_기초입력,
        tab_갑지,
        tab_내역서,
        tab_투입인원및내역,
        tab_산정기준,
        tab_노임단가,
        tab_손해보험요율,
    ) = st.tabs([
        "기초입력",
        "갑지",
        "내역서",
        "투입인원 및 내역",
        "투입인원수 산정기준",
        "노임단가",
        "보험요율",
    ])

    # ─── 기초입력 탭 ───
    with tab_기초입력:
        st.header("기초입력")

        # 1) 용역명_env
        st.text_input(
            "용역명",
            value=st.session_state.get("용역명_env", ""),
            key="용역명_env",
        )

        # 2) 발주기관명_env
        st.text_input(
            "발주기관명",
            value=st.session_state.get("발주기관명_env", ""),
            key="발주기관명_env",
        )

        # 3) 설계유형_env 선택
        env_options = ["소규모 환경영향평가 대행"]
        current_env = st.session_state.get("설계유형_env", env_options[0])
        idx_env = env_options.index(current_env) if current_env in env_options else 0
        st.radio(
            "설계유형을 선택하세요",
            env_options,
            index=idx_env,
            key="설계유형_env",
        )

        # 4) 대상 면적_env 입력
        st.number_input(
            "대상 면적 (㎡)",
            min_value=0.0,
            step=10.0,
            value=st.session_state.get("면적_env", 0.0),
            key="면적_env",
        )

        st.markdown("보정계수를 산정하기 위한 추가 질문")
        st.selectbox(
            "1. 동·식물상 조사 특성",
            ["생태*자연도 2등급 및 3등급 권역", "생태*자연도 1등급 권역 및 별도관리지역"],
            key="보정_동식물",
        )
        st.selectbox(
            "2. 자연경관심의 대상",
            ["대상", "미대상"],
            key="보정_자연연경관심의",
        )
        st.selectbox(
            "3. 건강영향평가 대상",
            ["대상", "미대상"],
            key="보정_건강영향평가",
        )
        st.selectbox(
            "4. 수질오염총량계획 대상",
            ["대상", "미대상"],
            key="보정_수질오염총량계획",
        )

    # ─── 갑지 탭 ───
    with tab_갑지:
        import datetime

        용역명_env = st.session_state.get("용역명_env", "")
        발주기관명_env = st.session_state.get("발주기관명_env", "")
        도급예정액_env = st.session_state.get("도급예정액_env", 0)

        today = datetime.date.today().strftime("%Y-%m-%d")
        st.markdown(f"##### 날짜: {today}")

        st.markdown(
            f"<h2 style='text-align:center;'>{용역명_env}</h2>",
            unsafe_allow_html=True,
        )

        if 도급예정액_env <= 0:
            st.info("먼저 ‘내역서’ 탭에서 **산출 완료✅** 버튼을 눌러 금액을 확정하세요.")
        else:
            표시_용역비_env = int(도급예정액_env // 1000) * 1000
            st.write(f"**용역비:** {표시_용역비_env:,.0f} 원")

        st.write(f"**발주기관:** {발주기관명_env}")

        if 도급예정액_env > 0:
            excel_buf = build_환경_excel("template_env.xlsx")
            if excel_buf is not None:
                st.download_button(
                    label="⬇️ 환경영향평가 내역서(Excel) 다운로드",
                    data=excel_buf,
                    file_name=f"{용역명_env}_환경영향평가_내역서.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
        else:
            st.caption("※ 산출 완료 후 버튼이 활성화됩니다.")

    # ─── 내역서 탭 ───
    with tab_내역서:
        st.header("내역서")
        st.caption("※ 각 숫자를 수정한 뒤 **Enter** 를 눌러야 계산이 반영됩니다.")

        직접인건비_env = st.session_state.get("직접인건비_env")
        if 직접인건비_env is None:
            st.warning("먼저 ‘투입인원 및 내역’ 탭에서 직접인건비를 계산해 주세요.")
        else:
            제경비율_env = st.number_input(
                "제경비율 (110~120%)",
                value=110.0,
                step=0.1,
                key="제경비율_env",
            )
            직접경비_env = st.number_input(
                "직접경비 금액 (원)",
                value=5_000_000,
                step=1_000,
                key="직접경비_env",
            )
            기술료율_env = st.number_input(
                "기술료율 (20~40%)",
                value=20.0,
                step=0.1,
                key="기술료율_env",
            )
            공제율_env = st.number_input(
                "손해공제비율 (%)",
                value=0.432,
                step=0.001,
                key="공제율_env",
            )
            부가세율_env = st.number_input(
                "부가가치세율 (%)",
                value=10.0,
                step=0.1,
                key="부가세율_env",
            )

            제경비_env = 직접인건비_env * 제경비율_env / 100
            기술료_env = (직접인건비_env + 제경비_env) * 기술료율_env / 100
            손해공제비_env = (
                (직접인건비_env + 제경비_env + 직접경비_env + 기술료_env)
                * 공제율_env
                / 100
            )
            부가세_env = (
                (직접인건비_env
                + 제경비_env
                + 직접경비_env
                + 기술료_env
                + 손해공제비_env)
                * 부가세율_env
                / 100
            )
            도급예정액_env = (
                직접인건비_env
                + 제경비_env
                + 직접경비_env
                + 기술료_env
                + 손해공제비_env
                + 부가세_env
            )

            if st.button("✅ 산출 완료"):
                st.session_state["도급예정액_env"] = 도급예정액_env
                st.success(f"도급예정액 {도급예정액_env:,.0f}원이 저장되었습니다.")
            else:
                st.info("▶️ 값이 맞다면 ‘✅ 산출 완료’ 버튼을 눌러 주세요.")

            rows_env = [
                {
                    "공종": "직접인건비",
                    "규격": "-",
                    "수량": "-",
                    "단위": "",
                    "총액": 직접인건비_env,
                    "노무비": 직접인건비_env,
                    "경비": "",
                    "비고": "",
                },
                {
                    "공종": "제경비",
                    "규격": "직접인건비×율",
                    "수량": "-",
                    "단위": "",
                    "총액": "",
                    "노무비": "",
                    "경비": 제경비_env,
                    "비고": f"{제경비율_env}%",
                },
                {
                    "공종": "직접경비",
                    "규격": "제출도서 인쇄",
                    "수량": 1,
                    "단위": "식",
                    "총액": "",
                    "노무비": "",
                    "경비": 직접경비_env,
                    "비고": "",
                },
                {
                    "공종": "기술료",
                    "규격": "(직접인건비+제경비)×율",
                    "수량": "-",
                    "단위": "",
                    "총액": "",
                    "노무비": "",
                    "경비": 기술료_env,
                    "비고": f"{기술료율_env}%",
                },
                {
                    "공종": "손해공제비",
                    "규격": "용역비×율",
                    "수량": "-",
                    "단위": "",
                    "총액": "",
                    "노무비": "",
                    "경비": 손해공제비_env,
                    "비고": f"{공제율_env}",
                },
                {
                    "공종": "부가가치세",
                    "규격": "합계×율",
                    "수량": "-",
                    "단위": "",
                    "총액": "",
                    "노무비": "",
                    "경비": 부가세_env,
                    "비고": f"{부가세율_env}%",
                },
                {
                    "공종": "도급예정액",
                    "규격": "",
                    "수량": "-",
                    "단위": "",
                    "총액": 도급예정액_env,
                    "노무비": "",
                    "경비": "",
                    "비고": "",
                },
            ]
            df_env = pd.DataFrame(rows_env)

            for c in ["총액", "노무비", "경비"]:
                df_env[c] = df_env[c].apply(
                    lambda x: f"{int(x):,}" if isinstance(x, (int, float)) else x
                )

            st.dataframe(df_env[[
                "공종", "규격", "수량", "단위", "총액", "노무비", "경비", "비고"
            ]])

            st.session_state["df_detail_env"] = df_env

    # ─── 투입인원 및 내역 탭 ───
    with tab_투입인원및내역:
        st.header("투입인원 및 내역")

        기준결과_env = st.session_state.get("기준결과_env")
        노임단가_df_env = st.session_state.get("최종_단가_env")

        if 기준결과_env is None or 노임단가_df_env is None:
            st.warning("먼저 ‘산정기준’ 탭과 ‘노임단가’ 탭을 완료해주세요.")
        else:
            결과표 = 기준결과_env.copy()
            결과표 = 결과표[
                결과표["단위"].astype(str).str.strip() != ""
            ].reset_index(drop=True)

            for 직급 in 직급리스트:
                결과표[직급] = pd.to_numeric(결과표[직급], errors="coerce").fillna(0.0)

            is_technical = 결과표[직급리스트].sum(axis=1) > 0

            기간값 = {}
            n = len(결과표)
            half = (n + 1) // 2
            left, right = st.columns(2)

            with left:
                for idx, row in 결과표.iloc[:half].iterrows():
                    if not is_technical.iloc[idx]:
                        기간값[idx] = 0
                        continue
                    업무 = row["업무구분"]
                    단위 = str(row["단위"]).strip()

                    if 단위 == "식":
                        기본값 = 1
                        라벨 = f"{업무} (식)"
                    else:
                        기본값 = 1
                        라벨 = f"{업무} 기간 (일)"

                    값 = st.number_input(
                        라벨,
                        min_value=0,
                        step=1,
                        value=int(st.session_state.get(f"기간_{idx}", 기본값)),
                        key=f"기간_L_{idx}"
                    )
                    기간값[idx] = 값

            with right:
                for idx, row in 결과표.iloc[half:].iterrows():
                    if not is_technical.iloc[idx]:
                        기간값[idx] = 0
                        continue
                    업무 = row["업무구분"]
                    단위 = str(row["단위"]).strip()

                    if 단위 == "식":
                        기본값 = 1
                        라벨 = f"{업무} (식)"
                    else:
                        기본값 = 1
                        라벨 = f"{업무} 기간 (일)"

                    값 = st.number_input(
                        라벨,
                        min_value=0,
                        step=1,
                        value=int(st.session_state.get(f"기간_{idx}", 기본값)),
                        key=f"기간_R_{idx}"
                    )
                    기간값[idx] = 값

            결과표["기간"] = [기간값[i] for i in range(n)]

            노임단가_df_env.columns = [c.strip() for c in 노임단가_df_env.columns]
            노임단가_df_env["직종명"] = 노임단가_df_env["직종명"].astype(str).str.strip()

            env_col = "환경"
            노임단가_df_env[env_col] = (
                노임단가_df_env[env_col]
                .astype(str)
                .str.replace(",", "")
                .str.strip()
                .astype(float)
            )

            단가사전 = {}
            for 직급 in 직급리스트:
                sub = 노임단가_df_env[노임단가_df_env["직종명"] == 직급]
                단가사전[직급] = float(sub[env_col].iloc[0]) if not sub.empty else 0.0

            계산된_계 = []
            for idx, row in 결과표.iterrows():
                if not is_technical.iloc[idx]:
                    계산된_계.append(0.0)
                    continue
                인건비합 = sum(row[직급] * 단가사전.get(직급, 0.0) for 직급 in 직급리스트)
                사전체 = round(인건비합 * row["기간"], 2)
                계산된_계.append(사전체)

            결과표["계"] = 계산된_계

            표시열 = ["업무구분", "계"] + 직급리스트 + ["기간"]

            sum_계_env = 결과표["계"].sum()
            총계행 = {c: "" for c in 표시열}
            총계행["업무구분"] = "총계"
            총계행["계"] = sum_계_env
            total_df = pd.DataFrame([총계행])

            final_df = pd.concat([total_df, 결과표[표시열]], ignore_index=True)

            def fmt(x):
                if isinstance(x, (int, float)):
                    return "" if x == 0 else f"{x:,.2f}"
                return x

            for c in ["계", "기간"] + 직급리스트:
                final_df[c] = final_df[c].apply(fmt)

            st.session_state["직접인건비_env"] = sum_계_env
            st.session_state["투입인원DF_env"] = final_df

            st.subheader("📊 기술자별 투입 인원 및 총액")
            st.dataframe(final_df)

    # ─── 산정기준 탭 ───
    with tab_산정기준:
        st.header("투입인원수 산정기준")

        설계유형_env = st.session_state.get("설계유형_env")
        대상_면적_env = st.session_state.get("면적_env", 0)
        q1 = st.session_state.get("보정_동식물")
        q2 = st.session_state.get("보정_자연연경관심의")
        q3 = st.session_state.get("보정_건강영향평가")
        q4 = st.session_state.get("보정_수질오염총량계획")

        if 설계유형_env == "소규모 환경영향평가 대행":
            기준표_env = load_env_basis(설계유형_env).copy()
            A = float(대상_면적_env)

            α1_list = []
            for _, row in 기준표_env.iterrows():
                rule = row.get("환산계수", None)
                α1 = 1.00
                if rule == 1:
                    α1 = 1.00
                elif rule == 2:
                    if A <= 10000:
                        α1 = 0.25
                    elif A <= 100000:
                        α1 = round((A / 100000) ** 0.6, 3)
                    elif A <= 1000000:
                        α1 = round((A / 100000) ** 0.3, 3)
                    else:
                        α1 = 2.00
                elif rule == 3:
                    if A <= 30000:
                        α1 = 0.49
                    elif A <= 1000000:
                        α1 = round((A / 100000) ** 0.6, 3)
                    else:
                        α1 = 3.98
                else:
                    α1 = 1.00
                α1_list.append(α1)

            기준표_env["α₁(환산계수)"] = α1_list

            factor_가 = 1.00 if q1 == "생태*자연도 2등급 및 3등급 권역" else 1.20
            factor_나 = 1.00 if q2 == "미대상" else 2.00
            factor_다 = 1.00 if q3 == "미대상" else 1.15
            factor_라 = 1.00 if q4 == "미대상" else 1.40

            for 직급 in 직급리스트:
                계산값_list = []
                계산식_list = []
                for _, row in 기준표_env.iterrows():
                    base = row.get(직급, 0)
                    α1 = row["α₁(환산계수)"]
                    보정가 = row.get("보정계수(가)", "")
                    보정나 = row.get("보정계수(나)", "")
                    보정다 = row.get("보정계수(다)", "")
                    보정라 = row.get("보정계수(라)", "")

                    if not isinstance(base, (int, float)) or base <= 0:
                        계산값_list.append("")
                        계산식_list.append("")
                        continue

                    v = base * α1
                    parts = [f"{α1:.3f}"]

                    if 보정가 == "반영":
                        v *= factor_가
                        parts.append(f"{factor_가:.2f}")
                    if 보정나 == "반영":
                        v *= factor_나
                        parts.append(f"{factor_나:.2f}")
                    if 보정다 == "반영":
                        v *= factor_다
                        parts.append(f"{factor_다:.2f}")
                    if 보정라 == "반영":
                        v *= factor_라
                        parts.append(f"{factor_라:.2f}")

                    v_rounded = round(v, 2)
                    계산값_list.append(v_rounded)

                    formula = f"{base:.2f} × " + " × ".join(parts)
                    계산식_list.append(formula)

                기준표_env[직급] = 계산값_list
                기준표_env[f"{직급}_계산식"] = 계산식_list

            표시열 = ["업무구분", "단위"]
            for 직급 in 직급리스트:
                표시열 += [직급, f"{직급}_계산식"]

            df_display_env = 기준표_env[표시열].copy()
            df_display_env = df_display_env.replace(0, "").fillna("")

            st.subheader(f"📊 {설계유형_env} 기준표 (환산계수 + 보정계수 반영)")
            st.dataframe(df_display_env)

            st.session_state["기준결과_env"] = 기준표_env

        else:
            st.info("‘소규모 환경영향평가 대행’을 선택해야 산정기준이 표시됩니다.")

    # ─── 노임단가 탭 ───
    with tab_노임단가:
        st.header("노임단가")
        df_wage_env = load_노임단가()
        st.dataframe(df_wage_env)
        st.session_state["최종_단가_env"] = df_wage_env

    # ─── 보험요율 탭 ───
    with tab_손해보험요율:
        st.header("보험요율")
        df_ins_env = load_손해보험요율()
        st.dataframe(df_ins_env)
        st.session_state["보험요율DF_env"] = df_ins_env

def main():
    st.sidebar.header("1️⃣ 설계 분야 선택")
    option = st.sidebar.radio(
        "어떤 설계를 하시나요?",
        ("조경", "환경영향평가 대행")
    )
    handlers = {
        "조경": run_조경,
        "환경영향평가 대행": run_환경영향평가대행
    }
    handlers[option]()

if __name__ == "__main__":
    main()